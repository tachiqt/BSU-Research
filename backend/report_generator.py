import os
import copy
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell


def _template_path(project_root=None):
    if project_root is None:
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(project_root, 'img', 'RESEARCH AL_4th Quarter Monitoring Report_2025.xlsx')


PUB_SHEET = 'RES-Publication'
PUB_HEADER_ROW = 5   
PUB_CAMPUS_ROW = 6   
PUB_DATA_START = 13 
PUB_SIG_LABEL_ROW = 34
PUB_SIG_NAME_ROW = 37
PUB_SIG_TITLE_ROW = 38
PUB_SIG_OFFICE_ROW = 39
PUB_COLS = {
    'no': 1, 'title': 2, 'project': 3, 'authors': 4, 'college_campus': 5,
    'source_fund': 6, 'status': 7, 'sdg': 8, 'requested': 9, 'venue': 10, 'pub_type': 11
}

PRES_SHEET = 'RES-Presentation'
PRES_HEADER_ROW = 5
PRES_CAMPUS_ROW = 6
PRES_SIG_NAME_ROW = 22
PRES_SIG_TITLE_ROW = 23
PRES_SIG_OFFICE_ROW = 24

IP_SHEET = 'RES-IP Assets'
IP_HEADER_ROW = 5
IP_CAMPUS_ROW = 6
IP_SIG_NAME_ROW = 24
IP_SIG_TITLE_ROW = 25
IP_SIG_OFFICE_ROW = 26


def _get_merged_top_left_cell(ws, row_1based, col_1based):
    cell = ws.cell(row=row_1based, column=col_1based)
    if not isinstance(cell, MergedCell):
        return cell
    for r in ws.merged_cells.ranges:
        if r.min_row <= row_1based <= r.max_row and r.min_col <= col_1based <= r.max_col:
            return ws.cell(row=r.min_row, column=r.min_col)
    return cell


def _set_cell(ws, row_1based, col_1based, value):
    if value is None:
        return
    c = _get_merged_top_left_cell(ws, row_1based, col_1based)
    if isinstance(c, MergedCell):
        return
    c.value = value


def _clear_cell(ws, row_1based, col_1based):
    cell = ws.cell(row=row_1based, column=col_1based)
    if not isinstance(cell, MergedCell):
        cell.value = None
        return
    top_left = _get_merged_top_left_cell(ws, row_1based, col_1based)
    if isinstance(top_left, MergedCell):
        return
    if top_left.row == row_1based and top_left.column == col_1based:
        top_left.value = None


def _copy_row_style(ws, src_row, dst_row, col_min=1, col_max=11):
    try:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    except Exception:
        pass
    for col in range(col_min, col_max + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        try:
            dst._style = copy.copy(src._style)
            dst.number_format = src.number_format
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.protection = copy.copy(src.protection)
        except Exception:
            pass


def _fill_header_and_signatures(ws, fiscal_year, quarter, campus, signatures, header_row, campus_row, name_row, title_row, office_row):
    _set_cell(ws, header_row, 1, f'Fiscal Year: {fiscal_year} Quarter: {quarter}')
    _set_cell(ws, campus_row, 1, f'Constituent Campus: {campus}')


def fill_publication_sheet(ws, publications, fiscal_year, quarter, campus, signatures):
    capacity = PUB_SIG_LABEL_ROW - PUB_DATA_START
    extra = max(0, len(publications) - capacity)
    if extra:
        ws.insert_rows(PUB_SIG_LABEL_ROW, amount=extra)
        for r in range(PUB_SIG_LABEL_ROW, PUB_SIG_LABEL_ROW + extra):
            _copy_row_style(ws, PUB_DATA_START, r, col_min=1, col_max=11)

    sig_label_row = PUB_SIG_LABEL_ROW + extra
    sig_name_row = PUB_SIG_NAME_ROW + extra
    sig_title_row = PUB_SIG_TITLE_ROW + extra
    sig_office_row = PUB_SIG_OFFICE_ROW + extra

    _fill_header_and_signatures(
        ws, fiscal_year, quarter, campus, signatures,
        PUB_HEADER_ROW, PUB_CAMPUS_ROW,
        sig_name_row, sig_title_row, sig_office_row
    )

    for r in range(PUB_DATA_START, sig_label_row):
        for c in range(1, 12):
            _clear_cell(ws, r, c)

    for i, pub in enumerate(publications, start=1):
        row = PUB_DATA_START + i - 1
        _set_cell(ws, row, PUB_COLS['no'], i)
        _set_cell(ws, row, PUB_COLS['title'], pub.get('title') or '')
        _set_cell(ws, row, PUB_COLS['project'], pub.get('project_title') or 'N/A')
        _set_cell(ws, row, PUB_COLS['authors'], pub.get('authors') or '')
        _set_cell(ws, row, PUB_COLS['college_campus'], pub.get('college_campus') or 'Batangas State University')
        _set_cell(ws, row, PUB_COLS['source_fund'], pub.get('source_fund') or 'N/A')
        _set_cell(ws, row, PUB_COLS['status'], pub.get('status') or 'N/A')
        _set_cell(ws, row, PUB_COLS['sdg'], pub.get('sdg') or 'N/A')
        _set_cell(ws, row, PUB_COLS['requested'], pub.get('requested') or 'N/A')
        _set_cell(ws, row, PUB_COLS['venue'], pub.get('venue') or '')
        _set_cell(ws, row, PUB_COLS['pub_type'], pub.get('pub_type') or 'Publication')


def fill_presentation_sheet(ws, fiscal_year, quarter, campus, signatures):
    _fill_header_and_signatures(
        ws, fiscal_year, quarter, campus, signatures,
        PRES_HEADER_ROW, PRES_CAMPUS_ROW,
        PRES_SIG_NAME_ROW, PRES_SIG_TITLE_ROW, PRES_SIG_OFFICE_ROW
    )


def fill_ip_assets_sheet(ws, fiscal_year, quarter, campus, signatures):
    _fill_header_and_signatures(
        ws, fiscal_year, quarter, campus, signatures,
        IP_HEADER_ROW, IP_CAMPUS_ROW,
        IP_SIG_NAME_ROW, IP_SIG_TITLE_ROW, IP_SIG_OFFICE_ROW
    )


def publications_to_report_rows(publications):
    rows = []
    for p in publications:
        year = p.get('year')
        if isinstance(year, str) and '/' in year:
            year = year.split('/')[0] if year else None
        month = p.get('month')
        if month is not None and isinstance(month, str):
            try:
                month = int(month)
            except (ValueError, TypeError):
                month = None
        pub_type = p.get('pub_type') or ''
        if not pub_type or not isinstance(pub_type, str):
            pub_type = 'Journal' if (p.get('venue') and 'journal' in (p.get('venue') or '').lower()) else 'Conference Proceeding'
        link = p.get('link') or ''
        doi = (p.get('doi') or '').strip()
        if not link and doi:
            link = ('https://doi.org/' + doi) if not doi.startswith('http') else doi
        mov_link = link or p.get('moy') or ''
        rows.append({
            'no': len(rows) + 1,
            'title': p.get('title') or 'Untitled Publication',
            'project_title': 'N/A',
            'authors': p.get('authors') or '',
            'college_campus': p.get('college_campus') or 'Batangas State University',
            'venue': p.get('venue') or '',
            'pub_type': pub_type,
            'year': year,
            'month': month,
            'source_fund': p.get('source_fund') or 'Non-funded',
            'indexing': p.get('indexing') or 'Scopus',
            'publisher': p.get('publisher') or '',
            'mov_link': mov_link,
            'moy': mov_link,
        })
    return rows


def _normalize_pub_type(pub_type):
    if not pub_type:
        return 'Other Type'
    s = str(pub_type).strip().lower()
    if 'journal' in s:
        return 'Journal'
    if 'conference' in s or 'proceeding' in s:
        return 'Conference Proceeding'
    return 'Other Type'


def _quarter_from_month(month):
    if month is None:
        return 0
    try:
        m = int(month)
        if 1 <= m <= 3:
            return 1
        if 4 <= m <= 6:
            return 2
        if 7 <= m <= 9:
            return 3
        if 10 <= m <= 12:
            return 4
    except (ValueError, TypeError):
        pass
    return 0


REPORT_COLUMNS = [
    'No.', 'Article Title', 'Author/s', 'College, Campus', 'Type of Publication',
    'Source of Fund', 'Journal or Conference Proceeding Title', 'Indexing', 'Publisher', 'MOV'
]
QUARTER_NAMES = {1: 'FIRST QUARTER', 2: 'SECOND QUARTER', 3: 'THIRD QUARTER', 4: 'FOURTH QUARTER'}
TYPE_SHEET_NAMES = {'Journal': 'Journal', 'Conference Proceeding': 'Conference Proceeding', 'Other': 'Other Type'}

REPORT_COLUMN_WIDTHS = {
    1: 6,  
    2: 45,  
    3: 28,  
    4: 28, 
    5: 18,  
    6: 14, 
    7: 38,  
    8: 12, 
    9: 28,  
    10: 50, 
}
QUARTER_HEADER_START_COL = 5
QUARTER_HEADER_END_COL = 10


def _blue_fill():
    from openpyxl.styles import PatternFill
    return PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')


def _header_font():
    from openpyxl.styles import Font
    return Font(bold=True, color='FFFFFF', size=11)


def _thin_border():
    from openpyxl.styles import Border, Side
    thin = Side(style='thin', color='000000')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_report_column_widths(ws, num_cols=10):
    for col, width in REPORT_COLUMN_WIDTHS.items():
        if col <= num_cols:
            ws.column_dimensions[get_column_letter(col)].width = width


def _quarter_label_to_number(quarter):
    if not quarter:
        return None
    q = str(quarter).strip().lower()
    if q == 'all':
        return None
    return {'1st': 1, '2nd': 2, '3rd': 3, '4th': 4}.get(q)


def build_report_by_type_and_quarter(fiscal_year, campus, publications, signatures, project_root=None, force_quarter=None):
    from openpyxl.styles import Alignment, Font

    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    blue_fill = _blue_fill()
    header_font = _header_font()
    thin_border = _thin_border()
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    by_type = {}
    for p in publications:
        pt = _normalize_pub_type(p.get('pub_type'))
        if force_quarter is not None and 1 <= force_quarter <= 4:
            q = force_quarter
        else:
            q = _quarter_from_month(p.get('month')) or 4
        by_type.setdefault(pt, {}).setdefault(q, []).append(p)

    ws = wb.create_sheet(title='Publications')
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    _apply_report_column_widths(ws)
    row = 1

    for type_key in ('Journal', 'Conference Proceeding', 'Other Type'):
        quarters_data = by_type.get(type_key, {})
        total_entries_type = sum(len(quarters_data.get(q, [])) for q in (1, 2, 3, 4))
        
        if total_entries_type == 0:
            continue

        tc = ws.cell(row=row, column=1, value=type_key.upper())
        tc.fill = blue_fill
        tc.font = header_font
        tc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(REPORT_COLUMNS))
        row += 1

        for col, label in enumerate(REPORT_COLUMNS, start=1):
            c = ws.cell(row=row, column=col, value=label)
            c.fill = blue_fill
            c.font = header_font
            c.border = thin_border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row += 1

        global_no = 0 
        for qnum in (1, 2, 3, 4):
            entries = quarters_data.get(qnum, [])
            if not entries:
                continue
            
            section_name = QUARTER_NAMES.get(qnum, f'QUARTER {qnum}')
            for col in range(1, len(REPORT_COLUMNS) + 1):
                c = ws.cell(row=row, column=col)
                c.fill = blue_fill
                c.font = Font(bold=True, color='000000', size=11)
                c.border = thin_border
                c.alignment = wrap_alignment

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(REPORT_COLUMNS))
            c1 = ws.cell(row=row, column=1, value=section_name)
            c1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            row += 1

            for pub in entries:
                global_no += 1
                for col in range(1, len(REPORT_COLUMNS) + 1):
                    ws.cell(row=row, column=col).alignment = wrap_alignment
                    ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=1, value=global_no)
                ws.cell(row=row, column=2, value=pub.get('title') or '')
                ws.cell(row=row, column=3, value=pub.get('authors') or '')
                ws.cell(row=row, column=4, value=pub.get('college_campus') or 'Batangas State University')
                ws.cell(row=row, column=5, value=pub.get('pub_type') or type_key)
                ws.cell(row=row, column=6, value=pub.get('source_fund') or 'Non-funded')
                ws.cell(row=row, column=7, value=pub.get('venue') or '')
                ws.cell(row=row, column=8, value=pub.get('indexing') or 'Scopus')
                ws.cell(row=row, column=9, value=pub.get('publisher') or '')
                mov_url = pub.get('mov_link') or pub.get('moy') or pub.get('link') or ''
                if pub.get('doi') and not mov_url:
                    doi = (pub.get('doi') or '').strip()
                    mov_url = ('https://doi.org/' + doi) if doi and not doi.startswith('http') else doi
                mov_cell = ws.cell(row=row, column=10)
                mov_cell.value = mov_url or pub.get('title') or ''
                if mov_url and (mov_url.startswith('http://') or mov_url.startswith('https://')):
                    mov_cell.hyperlink = mov_url
                    mov_cell.font = Font(color='0563C1', underline='single')
                row += 1
        
        row += 1 

    if row == 1:
         ws.cell(row=1, column=1, value='No publications in the selected period.')

    return wb


def build_report(fiscal_year, quarter, campus, publications, signatures, project_root=None):
    force_q = _quarter_label_to_number(quarter)
    return build_report_by_type_and_quarter(
        str(fiscal_year), str(campus), publications, signatures, project_root, force_quarter=force_q
    )


def get_preview_data(publications):
    return publications_to_report_rows(publications)
