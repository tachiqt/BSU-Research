import pandas as pd
import os
from typing import List, Dict, Optional, Tuple

# Optional openpyxl for appending to Excel (used by append_faculty_to_excel)
try:
    from openpyxl import load_workbook
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

def load_faculty_from_db_or_excel(file_path: str = None, sheet_name: str = None, prefer_db: bool = True) -> List[Dict]:
    """
    Load faculty data from database (preferred) or Excel file (fallback)
    
    Args:
        file_path: Path to Excel file (only used if database is empty or prefer_db=False)
        sheet_name: Sheet name in Excel (only used if reading from Excel)
        prefer_db: If True, try database first, fallback to Excel if empty
    
    Returns:
        List of faculty dictionaries
    """
    if prefer_db:
        try:
            from database import load_faculty_from_db, get_faculty_count
            count = get_faculty_count()
            if count > 0:
                print(f"Loading {count} faculty members from database...")
                return load_faculty_from_db()
            else:
                print("Database is empty, falling back to Excel...")
        except Exception as e:
            print(f"Error loading from database: {e}, falling back to Excel...")
    
    # Fallback to Excel
    return load_faculty_from_excel(file_path, sheet_name)

def load_faculty_from_excel(file_path: str = None, sheet_name: str = None) -> List[Dict]:
    """
    Load faculty from Excel file. If sheet_name is None or empty, uses the first sheet.
    """
    if file_path is None:
        current_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(current_dir)
        file_path = os.path.join(project_root, 'img', 'ref.xlsx')
    
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        excel_file = pd.ExcelFile(file_path)
        available_sheets = excel_file.sheet_names
        
        if not available_sheets:
            raise ValueError(f"No sheets found in Excel file: {file_path}")
        
        # If no sheet name provided or empty, use first sheet
        if not sheet_name or sheet_name.strip() == '':
            matching_sheet = available_sheets[0]
            print(f"No sheet name specified, using first sheet: '{matching_sheet}'")
        else:
            # Try to find matching sheet (case-insensitive)
            matching_sheet = None
            sheet_name_lower = sheet_name.strip().lower()
            
            for sheet in available_sheets:
                if sheet.lower() == sheet_name_lower:
                    matching_sheet = sheet
                    break
            
            if not matching_sheet:
                # Try partial match
                for sheet in available_sheets:
                    if sheet_name_lower in sheet.lower() or sheet.lower() in sheet_name_lower:
                        matching_sheet = sheet
                        break
            
            if not matching_sheet:
                # Use first sheet as fallback
                print(f"Warning: Sheet '{sheet_name}' not found. Available sheets: {available_sheets}")
                print(f"Using first available sheet: '{available_sheets[0]}'")
                matching_sheet = available_sheets[0]
        df = pd.read_excel(file_path, sheet_name=matching_sheet)
        df.columns = df.columns.str.strip()
        name_col = None
        dept_col = None
        position_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'name' in col_lower and name_col is None:
                name_col = col
            if 'department' in col_lower and dept_col is None:
                dept_col = col
            if ('position' in col_lower or 'designation' in col_lower) and position_col is None:
                position_col = col
        
        if not name_col:
            raise ValueError(f"NAME column not found. Available columns: {list(df.columns)}")
        if not dept_col:
            raise ValueError(f"DEPARTMENT column not found. Available columns: {list(df.columns)}")
        
        faculty_list = []
        for _, row in df.iterrows():
            name = str(row[name_col]).strip()
            department = str(row[dept_col]).strip() if dept_col else ''
            position = str(row[position_col]).strip() if position_col else ''
            if name == 'nan' or not name or name == '':
                continue
            name_variants = _generate_name_variants(name)
            
            faculty_list.append({
                'name': name,
                'department': department if department != 'nan' else '',
                'position': position if position != 'nan' else '',
                'name_variants': name_variants,
                'original_name': name  
            })
        return faculty_list
    except Exception as e:
        raise Exception(f"Error reading Excel file: {str(e)}")


def _detect_excel_columns(ws) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    """
    Detect 1-based column indices for name, department, position from first row.
    Returns (name_col, dept_col, position_col).
    """
    name_col = dept_col = position_col = None
    for col_idx, cell in enumerate(ws[1], start=1):
        val = (cell.value or '').strip() if cell.value else ''
        val_lower = val.lower()
        if 'name' in val_lower and name_col is None:
            name_col = col_idx
        if 'department' in val_lower and dept_col is None:
            dept_col = col_idx
        if ('position' in val_lower or 'designation' in val_lower) and position_col is None:
            position_col = col_idx
    return (name_col, dept_col, position_col)


def append_faculty_to_excel(
    file_path: str,
    name: str,
    department: str,
    position: str = '',
    sheet_name: Optional[str] = None,
    skip_duplicate: bool = True,
) -> Dict:
    """
    Append a single faculty row to an existing Excel file.
    Uses openpyxl to preserve file format. Falls back to pandas if openpyxl fails.

    Returns:
        dict: {'success': True} or {'duplicate': True} if name already exists and skip_duplicate=True.
    Raises:
        Exception: If file not found, columns missing, or write fails.
    """
    name = (name or '').strip()
    department = (department or '').strip()
    position = (position or '').strip()
    if not name:
        raise ValueError("Name is required")
    if not department:
        raise ValueError("Department is required")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    if _OPENPYXL_AVAILABLE:
        wb = load_workbook(file_path, read_only=False)
        sheet_names = wb.sheetnames
        if sheet_name and sheet_name.strip():
            match = None
            for s in sheet_names:
                if s.strip().lower() == sheet_name.strip().lower():
                    match = s
                    break
            ws = wb[match] if match else wb[sheet_names[0]]
        else:
            ws = wb.active if wb.active else wb[sheet_names[0]]
        name_col, dept_col, position_col = _detect_excel_columns(ws)
        if not name_col or not dept_col:
            raise ValueError(
                "Excel must have columns containing 'Name' and 'Department'. "
                f"Found first row: {[ws.cell(1, c).value for c in range(1, ws.max_column + 1)]}"
            )
        max_row = ws.max_row
        for r in range(2, max_row + 1):
            existing_name = (ws.cell(row=r, column=name_col).value or '')
            if isinstance(existing_name, str) and existing_name.strip().lower() == name.lower():
                return {'duplicate': True}
        next_row = max_row + 1
        ws.cell(row=next_row, column=name_col, value=name)
        ws.cell(row=next_row, column=dept_col, value=department)
        if position_col:
            ws.cell(row=next_row, column=position_col, value=position)
        wb.save(file_path)
        return {'success': True}

    # Fallback: pandas read, append row, write back (overwrites file)
    excel_file = pd.ExcelFile(file_path)
    available_sheets = excel_file.sheet_names
    use_sheet = (sheet_name or '').strip() or available_sheets[0]
    for s in available_sheets:
        if s.strip().lower() == (sheet_name or '').strip().lower():
            use_sheet = s
            break
    df = pd.read_excel(file_path, sheet_name=use_sheet)
    df.columns = df.columns.str.strip()
    name_col = dept_col = position_col = None
    for col in df.columns:
        col_lower = col.lower()
        if 'name' in col_lower and name_col is None:
            name_col = col
        if 'department' in col_lower and dept_col is None:
            dept_col = col
        if ('position' in col_lower or 'designation' in col_lower) and position_col is None:
            position_col = col
    if not name_col or not dept_col:
        raise ValueError("Excel must have columns containing 'Name' and 'Department'.")
    if skip_duplicate and name_col in df.columns:
        if df[name_col].astype(str).str.strip().str.lower().eq(name.lower()).any():
            return {'duplicate': True}
    new_row = {c: '' for c in df.columns}
    new_row[name_col] = name
    new_row[dept_col] = department
    if position_col:
        new_row[position_col] = position
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    df.to_excel(file_path, sheet_name=use_sheet, index=False, engine='openpyxl')
    return {'success': True}


def _generate_name_variants(name: str) -> List[str]:
    variants = [name.strip()] 
    name = name.strip()
    if not name:
        return variants
    
    if ',' in name:
        parts = [p.strip() for p in name.split(',')]
        if len(parts) == 2:
            last_name = parts[0]
            first_middle = parts[1]
            variants.append(f"{first_middle} {last_name}")
            first_middle_parts = first_middle.split()
            if len(first_middle_parts) >= 1:
                first_name = first_middle_parts[0]
                variants.append(f"{first_name} {last_name}")
                if len(first_name) > 0:
                    variants.append(f"{first_name[0]}. {last_name}")
                if len(first_middle_parts) >= 2:
                    middle_initials = '. '.join([p[0] for p in first_middle_parts if p])
                    variants.append(f"{middle_initials}. {last_name}")
    variants = [v for v in variants if v and v.strip()]
    return list(set(variants))

def match_author_to_faculty(author_name: str, faculty_list: List[Dict]) -> Dict:
    if not author_name or not author_name.strip():
        return None
    
    author_name = author_name.strip().rstrip(',.').strip()
    author_lower = author_name.lower()
    scopus_last = None
    scopus_initials = None
    import re
    
    if ',' in author_name:
        format2_match = re.match(r'^([A-Za-z\s]+?)\s+([A-Z](?:\.[A-Z])*(?:\.[A-Z]+)?)\s*,?\s*$', author_name.strip())
        if format2_match:
            scopus_last = format2_match.group(1).strip()
            scopus_initials = format2_match.group(2).replace('.', '').replace(' ', '').strip()
        else:
            author_parts = [p.strip() for p in author_name.split(',')]
            if len(author_parts) >= 2:
                scopus_last = author_parts[0].strip()
                scopus_initials = author_parts[1].strip().replace('.', '').replace(',', '').strip()
            elif len(author_parts) == 1:
                scopus_last = author_parts[0].strip()
                scopus_initials = ''
    else:
        match = re.match(r'^([A-Za-z\s]+?)\s+([A-Z](?:\.[A-Z])*(?:\.[A-Z]+)?)$', author_name.strip())
        if match:
            scopus_last = match.group(1).strip()
            scopus_initials = match.group(2).replace('.', '').replace(' ', '').strip()
        else:
            name_parts = author_name.split()
            if len(name_parts) >= 2:
                scopus_last = name_parts[-1]
                scopus_initials = ''.join([p[0] for p in name_parts[:-1] if p and len(p) > 0])
            elif len(name_parts) == 1:
                scopus_last = name_parts[0]
                scopus_initials = ''
    
    if not scopus_last:
        return None
    
    best_match = None
    best_score = 0
    scopus_last_lower = scopus_last.lower() if scopus_last else ''
    scopus_initials_clean = scopus_initials.upper().replace('.', '').replace(' ', '') if scopus_initials else ''
    author_parts_all = author_name.split()
    if len(author_parts_all) >= 2 and not scopus_last:
        # Might be "First Last" format
        potential_last = author_parts_all[-1]
        potential_first = author_parts_all[0]
        scopus_last_lower = potential_last.lower()
        scopus_initials_clean = potential_first[0].upper() if potential_first else ''
    
    for faculty in faculty_list:
        for variant in faculty.get('name_variants', [faculty['name']]):
            variant_clean = variant.strip()
            variant_lower = variant_clean.lower()
            if variant_lower == author_lower:
                return faculty
            if ',' in variant_clean:
                variant_parts = [p.strip() for p in variant_clean.split(',')]
                if len(variant_parts) >= 2:
                    excel_last = variant_parts[0].strip()
                    excel_first_middle = variant_parts[1].strip()
                    excel_initials = ''.join([p[0].upper() for p in excel_first_middle.split() if p and len(p) > 0])
                    excel_initials = excel_initials.replace('.', '').replace(' ', '')
                    if excel_last.lower() == scopus_last_lower:
                        if scopus_initials_clean and excel_initials:
                            scopus_normalized = scopus_initials_clean.upper().replace('.', '').replace(' ', '')
                            excel_normalized = excel_initials.upper().replace('.', '').replace(' ', '')
                            if scopus_normalized == excel_normalized:
                                return faculty  
                            if len(scopus_normalized) == len(excel_normalized):
                                if scopus_normalized == excel_normalized:
                                    return faculty
                            if len(scopus_normalized) > 0 and len(excel_normalized) > 0:
                                if scopus_normalized[0] == excel_normalized[0]:
                                    if len(scopus_normalized) == len(excel_normalized):
                                        score = 0.9  
                                    else:
                                        score = 0.8  
                                    if score > best_score:
                                        best_score = score
                                        best_match = faculty
                        else:
                            score = 0.7
                            if score > best_score:
                                best_score = score
                                best_match = faculty
            else:
                variant_parts = variant_clean.split()
                if len(variant_parts) >= 2:
                    variant_last = variant_parts[-1].lower()
                    if variant_last == scopus_last_lower:
                        score = 0.7
                        if score > best_score:
                            best_score = score
                            best_match = faculty
    
    if best_score >= 0.7:
        return best_match
    
    return None
